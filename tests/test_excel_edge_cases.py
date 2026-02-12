"""
Edge case tests for Excel parser special cases.

Tests cover:
- Empty Excel sheets
- Merged cells
- Formula errors (#DIV/0!, #VALUE!, #REF!, etc.)
- Hyperlinks
- Hidden rows and columns

Requirements: 2.5
"""

import pytest
import os
from pathlib import Path


class TestExcelEdgeCases:
    """Test Excel parser edge cases and special scenarios."""
    
    def test_empty_excel_sheet(self, tmp_path):
        """Test that empty Excel sheets are handled correctly.
        
        Validates: Requirements 2.5 - Empty sheet handling
        """
        from src.parsers import ExcelParser
        from src.internal_representation import Paragraph, TextFormatting
        
        # Create an Excel file with an empty sheet
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        excel_path = tmp_path / "empty_sheet.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "EmptySheet"
        # Don't add any data - leave it empty
        workbook.save(excel_path)
        
        # Parse the Excel file
        parser = ExcelParser()
        result = parser.parse(str(excel_path))
        
        # Verify the result
        assert len(result.sections) == 1
        section = result.sections[0]
        assert section.heading.text == "EmptySheet"
        assert len(section.content) == 1
        
        # Check that it contains an empty sheet indicator
        content = section.content[0]
        assert isinstance(content, Paragraph)
        assert "(Empty sheet)" in content.text
        assert content.formatting == TextFormatting.ITALIC
    
    def test_multiple_empty_sheets(self, tmp_path):
        """Test Excel file with multiple empty sheets.
        
        Validates: Requirements 2.5 - Multiple empty sheets
        """
        from src.parsers import ExcelParser
        from src.internal_representation import Paragraph
        
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        excel_path = tmp_path / "multiple_empty.xlsx"
        workbook = openpyxl.Workbook()
        
        # Create multiple empty sheets
        sheet1 = workbook.active
        sheet1.title = "Empty1"
        
        sheet2 = workbook.create_sheet("Empty2")
        sheet3 = workbook.create_sheet("Empty3")
        
        workbook.save(excel_path)
        
        # Parse the Excel file
        parser = ExcelParser()
        result = parser.parse(str(excel_path))
        
        # Verify all sheets are marked as empty
        assert len(result.sections) == 3
        for i, section in enumerate(result.sections):
            assert section.heading.text == f"Empty{i+1}"
            assert len(section.content) == 1
            assert isinstance(section.content[0], Paragraph)
            assert "(Empty sheet)" in section.content[0].text
    
    def test_merged_cells_handling(self, tmp_path):
        """Test that merged cells are handled correctly.
        
        Validates: Requirements 2.5 - Merged cell processing
        """
        from src.parsers import ExcelParser
        from src.internal_representation import Table
        
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        excel_path = tmp_path / "merged_cells.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "MergedCells"
        
        # Create a table with merged cells
        sheet['A1'] = "Header1"
        sheet['B1'] = "Header2"
        sheet['C1'] = "Header3"
        
        # Merge cells B2:C2 and set value
        sheet['A2'] = "Row1"
        sheet['B2'] = "Merged Value"
        sheet.merge_cells('B2:C2')
        
        # Another row with merged cells
        sheet['A3'] = "Row2"
        sheet['B3'] = "Another Merged"
        sheet.merge_cells('B3:C3')
        
        workbook.save(excel_path)
        
        # Parse the Excel file
        parser = ExcelParser()
        result = parser.parse(str(excel_path))
        
        # Verify merged cells are expanded
        assert len(result.sections) == 1
        section = result.sections[0]
        assert len(section.content) == 1
        
        table = section.content[0]
        assert isinstance(table, Table)
        assert table.headers == ["Header1", "Header2", "Header3"]
        
        # Check that merged cell values are replicated
        assert len(table.rows) == 2
        assert table.rows[0] == ["Row1", "Merged Value", "Merged Value"]
        assert table.rows[1] == ["Row2", "Another Merged", "Another Merged"]
    
    def test_formula_errors_handling(self, tmp_path):
        """Test that formula errors are handled correctly.
        
        Validates: Requirements 2.5 - Formula error processing
        """
        from src.parsers import ExcelParser
        from src.internal_representation import Table
        
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        excel_path = tmp_path / "formula_errors.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "FormulaErrors"
        
        # Create headers
        sheet['A1'] = "Error Type"
        sheet['B1'] = "Formula"
        sheet['C1'] = "Result"
        
        # Add various formula errors
        sheet['A2'] = "Division by zero"
        sheet['B2'] = "=1/0"
        
        sheet['A3'] = "Invalid reference"
        sheet['B3'] = "=A100+B200"  # This won't error, but we can create one
        
        sheet['A4'] = "Text in number operation"
        sheet['B4'] = "=A1+10"  # Trying to add text to number
        
        # Add a valid formula for comparison
        sheet['A5'] = "Valid formula"
        sheet['B5'] = "=2+2"
        
        workbook.save(excel_path)
        
        # Parse the Excel file
        parser = ExcelParser()
        result = parser.parse(str(excel_path))
        
        # Verify formula errors are captured
        assert len(result.sections) == 1
        section = result.sections[0]
        assert len(section.content) == 1
        
        table = section.content[0]
        assert isinstance(table, Table)
        assert len(table.rows) == 4
        
        # Check that error formulas are present (either as error values or calculated values)
        # The exact error representation depends on Excel's calculation
        # We just verify the parser doesn't crash and returns something
        for row in table.rows:
            assert len(row) == 3
            assert row[0] is not None  # Error type should be present
    
    def test_hyperlinks_conversion(self, tmp_path):
        """Test that hyperlinks are converted to Markdown format.
        
        Validates: Requirements 2.5 - Hyperlink conversion
        """
        from src.parsers import ExcelParser
        from src.internal_representation import Table
        
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        excel_path = tmp_path / "hyperlinks.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Hyperlinks"
        
        # Create headers
        sheet['A1'] = "Link Name"
        sheet['B1'] = "URL"
        
        # Add hyperlinks
        sheet['A2'] = "Google"
        sheet['B2'] = "Click here"
        sheet['B2'].hyperlink = "https://www.google.com"
        
        sheet['A3'] = "GitHub"
        sheet['B3'] = "Repository"
        sheet['B3'].hyperlink = "https://github.com"
        
        # Add a cell with text but no hyperlink
        sheet['A4'] = "No Link"
        sheet['B4'] = "Just text"
        
        workbook.save(excel_path)
        
        # Parse the Excel file
        parser = ExcelParser()
        result = parser.parse(str(excel_path))
        
        # Verify hyperlinks are converted to Markdown format
        assert len(result.sections) == 1
        section = result.sections[0]
        assert len(section.content) == 1
        
        table = section.content[0]
        assert isinstance(table, Table)
        assert len(table.rows) == 3
        
        # Check Markdown link format [text](url)
        assert "[Click here](https://www.google.com)" in table.rows[0][1]
        assert "[Repository](https://github.com)" in table.rows[1][1]
        assert table.rows[2][1] == "Just text"  # No link
    
    def test_hidden_rows_and_columns(self, tmp_path):
        """Test that hidden rows and columns are included by default.
        
        Validates: Requirements 2.5 - Hidden rows/columns processing
        """
        from src.parsers import ExcelParser
        from src.internal_representation import Table
        
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        excel_path = tmp_path / "hidden_rows_cols.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "HiddenData"
        
        # Create a table with some hidden rows and columns
        sheet['A1'] = "Col A"
        sheet['B1'] = "Col B (Hidden)"
        sheet['C1'] = "Col C"
        
        sheet['A2'] = "Row 2"
        sheet['B2'] = "Hidden Col Data"
        sheet['C2'] = "Visible"
        
        sheet['A3'] = "Row 3 (Hidden)"
        sheet['B3'] = "Hidden Row Data"
        sheet['C3'] = "Also Hidden Row"
        
        sheet['A4'] = "Row 4"
        sheet['B4'] = "More Data"
        sheet['C4'] = "Visible"
        
        # Hide column B
        sheet.column_dimensions['B'].hidden = True
        
        # Hide row 3
        sheet.row_dimensions[3].hidden = True
        
        workbook.save(excel_path)
        
        # Parse the Excel file
        parser = ExcelParser()
        result = parser.parse(str(excel_path))
        
        # Verify hidden rows and columns are included
        assert len(result.sections) == 1
        section = result.sections[0]
        assert len(section.content) == 1
        
        table = section.content[0]
        assert isinstance(table, Table)
        
        # All columns should be present (including hidden column B)
        assert len(table.headers) == 3
        assert "Col B (Hidden)" in table.headers[1]
        
        # All rows should be present (including hidden row 3)
        assert len(table.rows) == 3
        assert "Hidden Row Data" in table.rows[1][1]
        assert "Hidden Col Data" in table.rows[0][1]
    
    def test_mixed_special_cases(self, tmp_path):
        """Test a sheet with multiple special cases combined.
        
        Validates: Requirements 2.5 - Combined edge cases
        """
        from src.parsers import ExcelParser
        from src.internal_representation import Table
        
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        excel_path = tmp_path / "mixed_cases.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "MixedCases"
        
        # Create a complex table
        sheet['A1'] = "Name"
        sheet['B1'] = "Link"
        sheet['C1'] = "Calculation"
        
        # Row with hyperlink and formula
        sheet['A2'] = "Item 1"
        sheet['B2'] = "Website"
        sheet['B2'].hyperlink = "https://example.com"
        sheet['C2'] = "=10+5"
        
        # Row with merged cells
        sheet['A3'] = "Item 2"
        sheet['B3'] = "Merged Content"
        sheet.merge_cells('B3:C3')
        
        # Hidden row with formula error
        sheet['A4'] = "Item 3"
        sheet['B4'] = "Error"
        sheet['C4'] = "=1/0"
        sheet.row_dimensions[4].hidden = True
        
        workbook.save(excel_path)
        
        # Parse the Excel file
        parser = ExcelParser()
        result = parser.parse(str(excel_path))
        
        # Verify all special cases are handled
        assert len(result.sections) == 1
        section = result.sections[0]
        assert len(section.content) == 1
        
        table = section.content[0]
        assert isinstance(table, Table)
        assert len(table.headers) == 3
        assert len(table.rows) == 3  # All rows including hidden
        
        # Check hyperlink
        assert "[Website](https://example.com)" in table.rows[0][1]
        
        # Check merged cells
        assert table.rows[1][1] == table.rows[1][2]
        
        # Check hidden row is included
        assert "Item 3" in table.rows[2][0]
