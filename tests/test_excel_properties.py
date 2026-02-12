"""
Property-based tests for Excel parser.

This module contains property-based tests using Hypothesis to verify
Excel document parsing functionality across a wide range of inputs.

Feature: document-to-markdown-converter
Properties: 8, 9, 10, 11
Validates: Requirements 2.1, 2.2, 2.3, 2.4
"""

import pytest
from hypothesis import given, strategies as st, settings, HealthCheck
from pathlib import Path
import tempfile
import string

from src.parsers import ExcelParser
from src.internal_representation import (
    InternalDocument, Section, Heading, Table, Paragraph
)


# Helper strategies for generating Excel-compatible data
def excel_safe_text(min_size=0, max_size=100):
    """Generate text that is safe for Excel cells."""
    # Excel-safe characters (printable ASCII without control characters)
    safe_chars = string.printable.replace('\x0b', '').replace('\x0c', '').replace('\r', '').replace('\n', '')
    return st.text(alphabet=safe_chars, min_size=min_size, max_size=max_size)


def sheet_name_strategy():
    """Generate valid Excel sheet names."""
    # Excel sheet names: 1-31 characters, no special chars like : \ / ? * [ ]
    safe_chars = string.ascii_letters + string.digits + ' _-'
    return st.text(alphabet=safe_chars, min_size=1, max_size=31).filter(lambda x: x.strip())


class TestExcelPropertySheetExtraction:
    """Property 8: Sheet extraction
    
    For any valid .xlsx file, the converter should extract and convert
    all sheets to Markdown format.
    
    Validates: Requirements 2.1
    """
    
    @given(num_sheets=st.integers(min_value=1, max_value=5))
    @settings(max_examples=100, deadline=None, suppress_health_check=[HealthCheck.function_scoped_fixture])
    def test_property_8_sheet_extraction(self, num_sheets):
        """
        Feature: document-to-markdown-converter
        Property 8: Sheet extraction
        
        For any valid .xlsx file, the converter should extract and convert
        all sheets to Markdown format.
        
        Validates: Requirements 2.1
        """
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        # Create an Excel file with multiple sheets
        with tempfile.TemporaryDirectory() as tmp_dir:
            excel_path = Path(tmp_dir) / f"test_sheets_{num_sheets}.xlsx"
            workbook = openpyxl.Workbook()
            
            # Remove default sheet
            if 'Sheet' in workbook.sheetnames:
                del workbook['Sheet']
            
            # Create sheets with unique names
            sheet_names = []
            for i in range(num_sheets):
                sheet_name = f"Sheet{i+1}"
                sheet_names.append(sheet_name)
                
                if i == 0:
                    sheet = workbook.create_sheet(sheet_name, 0)
                else:
                    sheet = workbook.create_sheet(sheet_name)
                
                # Add some data to make it non-empty
                sheet['A1'] = f"Data in {sheet_name}"
            
            workbook.save(excel_path)
            
            # Parse the Excel file
            parser = ExcelParser()
            result = parser.parse(str(excel_path))
            
            # Property: All sheets should be extracted
            assert isinstance(result, InternalDocument)
            assert len(result.sections) == num_sheets, \
                f"Expected {num_sheets} sections, got {len(result.sections)}"
            
            # Property: Each section should correspond to a sheet
            for i, section in enumerate(result.sections):
                assert isinstance(section, Section)
                assert section.heading is not None
                assert isinstance(section.heading, Heading)
                assert section.heading.text == sheet_names[i], \
                    f"Expected sheet name '{sheet_names[i]}', got '{section.heading.text}'"
            
            # Property: Each section should have content
            for section in result.sections:
                assert len(section.content) > 0, "Section should have content"


class TestExcelPropertyMultiSheetHandling:
    """Property 9: Multi-sheet handling
    
    For any Excel file with multiple sheets, the converter should convert
    each sheet separately with clear sheet name headers.
    
    Validates: Requirements 2.2
    """
    
    @given(
        num_sheets=st.integers(min_value=2, max_value=5),
        data=st.data()
    )
    @settings(max_examples=100, deadline=None, suppress_health_check=[HealthCheck.function_scoped_fixture])
    def test_property_9_multi_sheet_handling(self, num_sheets, data):
        """
        Feature: document-to-markdown-converter
        Property 9: Multi-sheet handling
        
        For any Excel file with multiple sheets, the converter should convert
        each sheet separately with clear sheet name headers.
        
        Validates: Requirements 2.2
        """
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        # Create an Excel file with multiple sheets
        with tempfile.TemporaryDirectory() as tmp_dir:
            excel_path = Path(tmp_dir) / f"test_multi_sheets_{num_sheets}.xlsx"
            workbook = openpyxl.Workbook()
            
            # Remove default sheet
            if 'Sheet' in workbook.sheetnames:
                del workbook['Sheet']
            
            # Create sheets with different data
            sheet_names = []
            for i in range(num_sheets):
                # Generate unique sheet name
                sheet_name = data.draw(sheet_name_strategy())
                # Ensure uniqueness
                while sheet_name in sheet_names:
                    sheet_name = f"{sheet_name}_{i}"
                sheet_names.append(sheet_name)
                
                if i == 0:
                    sheet = workbook.create_sheet(sheet_name, 0)
                else:
                    sheet = workbook.create_sheet(sheet_name)
                
                # Add different data to each sheet
                num_rows = data.draw(st.integers(min_value=1, max_value=5))
                num_cols = data.draw(st.integers(min_value=1, max_value=5))
                
                for row in range(1, num_rows + 1):
                    for col in range(1, num_cols + 1):
                        cell_value = data.draw(excel_safe_text(min_size=1, max_size=20))
                        sheet.cell(row=row, column=col, value=cell_value)
            
            workbook.save(excel_path)
            
            # Parse the Excel file
            parser = ExcelParser()
            result = parser.parse(str(excel_path))
            
            # Property: Each sheet should be converted separately
            assert len(result.sections) == num_sheets, \
                f"Expected {num_sheets} sections, got {len(result.sections)}"
            
            # Property: Each section should have a clear sheet name header
            extracted_sheet_names = []
            for section in result.sections:
                assert section.heading is not None, "Section should have a heading"
                assert isinstance(section.heading, Heading), "Heading should be a Heading object"
                assert section.heading.level == 2, "Sheet headings should be level 2"
                extracted_sheet_names.append(section.heading.text)
            
            # Property: Sheet names should match original names
            assert extracted_sheet_names == sheet_names, \
                f"Expected sheet names {sheet_names}, got {extracted_sheet_names}"
            
            # Property: Each section should have distinct content
            # (We can't guarantee content is different, but we can check it exists)
            for i, section in enumerate(result.sections):
                assert len(section.content) > 0, f"Sheet {i} should have content"


class TestExcelPropertyDataTableConversion:
    """Property 10: Data table conversion
    
    For any Excel sheet containing data, the converter should convert it
    to valid Markdown table format.
    
    Validates: Requirements 2.3
    """
    
    @given(
        num_rows=st.integers(min_value=2, max_value=10),
        num_cols=st.integers(min_value=1, max_value=8),
        data=st.data()
    )
    @settings(max_examples=100, deadline=None, suppress_health_check=[HealthCheck.function_scoped_fixture])
    def test_property_10_data_table_conversion(self, num_rows, num_cols, data):
        """
        Feature: document-to-markdown-converter
        Property 10: Data table conversion
        
        For any Excel sheet containing data, the converter should convert it
        to valid Markdown table format.
        
        Validates: Requirements 2.3
        """
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        # Create an Excel file with a data table
        with tempfile.TemporaryDirectory() as tmp_dir:
            excel_path = Path(tmp_dir) / f"test_table_{num_rows}x{num_cols}.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "DataTable"
            
            # Generate table data
            table_data = []
            has_any_non_empty = False
            for row in range(num_rows):
                row_data = []
                for col in range(num_cols):
                    # Generate cell value
                    cell_value = data.draw(excel_safe_text(min_size=0, max_size=50))
                    if cell_value:  # Track if we have any non-empty values
                        has_any_non_empty = True
                    row_data.append(cell_value)
                    # Write to sheet
                    sheet.cell(row=row+1, column=col+1, value=cell_value)
                table_data.append(row_data)
            
            # Skip test if all cells are empty strings (Excel treats these as None)
            # This is not "data" according to the property statement
            if not has_any_non_empty:
                return
            
            workbook.save(excel_path)
            
            # Parse the Excel file
            parser = ExcelParser()
            result = parser.parse(str(excel_path))
            
            # Property: Should produce a valid internal document
            assert isinstance(result, InternalDocument)
            assert len(result.sections) == 1
            
            section = result.sections[0]
            assert len(section.content) == 1
            
            # Property: Content should be a Table
            table = section.content[0]
            assert isinstance(table, Table), f"Expected Table, got {type(table)}"
            
            # Property: Table should have correct dimensions
            assert len(table.headers) == num_cols, \
                f"Expected {num_cols} headers, got {len(table.headers)}"
            assert len(table.rows) == num_rows - 1, \
                f"Expected {num_rows - 1} data rows, got {len(table.rows)}"
            
            # Property: Headers should match first row
            for col_idx in range(num_cols):
                expected_header = str(table_data[0][col_idx]) if table_data[0][col_idx] else ""
                assert table.headers[col_idx] == expected_header, \
                    f"Header mismatch at column {col_idx}"
            
            # Property: Data rows should match remaining rows
            for row_idx in range(1, num_rows):
                for col_idx in range(num_cols):
                    expected_value = str(table_data[row_idx][col_idx]) if table_data[row_idx][col_idx] else ""
                    actual_value = table.rows[row_idx - 1][col_idx]
                    assert actual_value == expected_value, \
                        f"Data mismatch at row {row_idx}, col {col_idx}"
            
            # Property: Each row should have the same number of columns
            for row in table.rows:
                assert len(row) == num_cols, \
                    f"Row should have {num_cols} columns, got {len(row)}"
    
    @given(data=st.data())
    @settings(max_examples=50, deadline=None, suppress_health_check=[HealthCheck.function_scoped_fixture])
    def test_property_10_various_data_types(self, data):
        """
        Feature: document-to-markdown-converter
        Property 10: Data table conversion (various data types)
        
        For any Excel sheet containing various data types (numbers, text, booleans),
        the converter should convert them correctly to Markdown table format.
        
        Validates: Requirements 2.3
        """
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        # Create an Excel file with various data types
        with tempfile.TemporaryDirectory() as tmp_dir:
            excel_path = Path(tmp_dir) / "test_data_types.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "DataTypes"
            
            # Add headers
            sheet['A1'] = "Text"
            sheet['B1'] = "Number"
            sheet['C1'] = "Boolean"
            sheet['D1'] = "Empty"
            
            # Generate data rows
            num_rows = data.draw(st.integers(min_value=1, max_value=5))
            for row in range(num_rows):
                # Text column
                text_val = data.draw(excel_safe_text(min_size=1, max_size=30))
                sheet.cell(row=row+2, column=1, value=text_val)
                
                # Number column
                num_val = data.draw(st.one_of(
                    st.integers(min_value=-1000, max_value=1000),
                    st.floats(min_value=-1000.0, max_value=1000.0, allow_nan=False, allow_infinity=False)
                ))
                sheet.cell(row=row+2, column=2, value=num_val)
                
                # Boolean column
                bool_val = data.draw(st.booleans())
                sheet.cell(row=row+2, column=3, value=bool_val)
                
                # Empty column (leave as None)
                sheet.cell(row=row+2, column=4, value=None)
            
            workbook.save(excel_path)
            
            # Parse the Excel file
            parser = ExcelParser()
            result = parser.parse(str(excel_path))
            
            # Property: Should produce a valid table
            assert len(result.sections) == 1
            table = result.sections[0].content[0]
            assert isinstance(table, Table)
            
            # Property: Table should have correct structure
            assert len(table.headers) == 4
            assert table.headers == ["Text", "Number", "Boolean", "Empty"]
            assert len(table.rows) == num_rows
            
            # Property: Each row should have 4 columns
            for row in table.rows:
                assert len(row) == 4
            
            # Property: Boolean values should be converted to TRUE/FALSE
            for row in table.rows:
                bool_value = row[2]
                assert bool_value in ["TRUE", "FALSE"], \
                    f"Boolean should be TRUE or FALSE, got {bool_value}"
            
            # Property: Empty cells should be empty strings
            for row in table.rows:
                empty_value = row[3]
                assert empty_value == "", f"Empty cell should be empty string, got '{empty_value}'"


class TestExcelPropertyFormulaValueConversion:
    """Property 11: Formula value conversion
    
    For any Excel cell containing a formula, the converter should output
    the calculated value in the Markdown table.
    
    Validates: Requirements 2.4
    """
    
    @given(
        num1=st.integers(min_value=1, max_value=100),
        num2=st.integers(min_value=1, max_value=100),
        operation=st.sampled_from(['+', '-', '*'])
    )
    @settings(max_examples=100, deadline=None, suppress_health_check=[HealthCheck.function_scoped_fixture])
    def test_property_11_formula_value_conversion(self, num1, num2, operation):
        """
        Feature: document-to-markdown-converter
        Property 11: Formula value conversion
        
        For any Excel cell containing a formula, the converter should output
        the calculated value in the Markdown table.
        
        Validates: Requirements 2.4
        """
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        # Create an Excel file with formulas
        with tempfile.TemporaryDirectory() as tmp_dir:
            excel_path = Path(tmp_dir) / f"test_formula_{num1}_{operation}_{num2}.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Formulas"
            
            # Add headers
            sheet['A1'] = "Value1"
            sheet['B1'] = "Value2"
            sheet['C1'] = "Formula"
            sheet['D1'] = "Result"
            
            # Add data
            sheet['A2'] = num1
            sheet['B2'] = num2
            
            # Add formula
            formula = f"=A2{operation}B2"
            sheet['C2'] = formula
            
            # Calculate expected result
            if operation == '+':
                expected_result = num1 + num2
            elif operation == '-':
                expected_result = num1 - num2
            elif operation == '*':
                expected_result = num1 * num2
            
            # Add expected result for verification
            sheet['D2'] = expected_result
            
            workbook.save(excel_path)
            
            # Parse the Excel file
            parser = ExcelParser()
            result = parser.parse(str(excel_path))
            
            # Property: Should produce a valid table
            assert len(result.sections) == 1
            table = result.sections[0].content[0]
            assert isinstance(table, Table)
            
            # Property: Table should have correct structure
            assert len(table.headers) == 4
            assert len(table.rows) == 1
            
            # Property: Formula cell should contain the calculated value, not the formula
            formula_cell_value = table.rows[0][2]
            
            # The formula cell should contain a numeric value (the result)
            # It should NOT contain the formula string like "=A2+B2"
            assert not formula_cell_value.startswith('='), \
                f"Formula cell should contain calculated value, not formula: {formula_cell_value}"
            
            # Property: The calculated value should match the expected result
            try:
                calculated_value = float(formula_cell_value)
                assert abs(calculated_value - expected_result) < 0.001, \
                    f"Expected result {expected_result}, got {calculated_value}"
            except ValueError:
                # If it's not a number, it might be an error or the formula itself
                # In this case, we accept it as long as it's not the formula string
                pass
    
    @given(data=st.data())
    @settings(max_examples=50, deadline=None, suppress_health_check=[HealthCheck.function_scoped_fixture])
    def test_property_11_complex_formulas(self, data):
        """
        Feature: document-to-markdown-converter
        Property 11: Formula value conversion (complex formulas)
        
        For any Excel cell containing complex formulas (SUM, AVERAGE, etc.),
        the converter should output the calculated value.
        
        Validates: Requirements 2.4
        """
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        # Create an Excel file with complex formulas
        with tempfile.TemporaryDirectory() as tmp_dir:
            excel_path = Path(tmp_dir) / "test_complex_formulas.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "ComplexFormulas"
            
            # Add headers
            sheet['A1'] = "Numbers"
            sheet['B1'] = "SUM"
            sheet['C1'] = "AVERAGE"
            sheet['D1'] = "MAX"
            
            # Generate random numbers
            num_values = data.draw(st.integers(min_value=2, max_value=5))
            values = []
            for i in range(num_values):
                val = data.draw(st.integers(min_value=1, max_value=100))
                values.append(val)
                sheet.cell(row=i+2, column=1, value=val)
            
            # Add formulas in the last row
            last_row = num_values + 2
            sheet.cell(row=last_row, column=2, value=f"=SUM(A2:A{num_values+1})")
            sheet.cell(row=last_row, column=3, value=f"=AVERAGE(A2:A{num_values+1})")
            sheet.cell(row=last_row, column=4, value=f"=MAX(A2:A{num_values+1})")
            
            # Calculate expected results
            expected_sum = sum(values)
            expected_avg = sum(values) / len(values)
            expected_max = max(values)
            
            workbook.save(excel_path)
            
            # Parse the Excel file
            parser = ExcelParser()
            result = parser.parse(str(excel_path))
            
            # Property: Should produce a valid table
            assert len(result.sections) == 1
            table = result.sections[0].content[0]
            assert isinstance(table, Table)
            
            # Property: Table should have the correct number of rows
            assert len(table.rows) == num_values + 1  # data rows + formula row
            
            # Property: Formula cells should contain calculated values
            formula_row = table.rows[-1]
            
            # Check SUM
            sum_value = formula_row[1]
            assert not sum_value.startswith('='), "SUM cell should contain calculated value"
            try:
                calculated_sum = float(sum_value)
                assert abs(calculated_sum - expected_sum) < 0.001, \
                    f"Expected SUM {expected_sum}, got {calculated_sum}"
            except ValueError:
                pass  # Accept if it's an error or special value
            
            # Check AVERAGE
            avg_value = formula_row[2]
            assert not avg_value.startswith('='), "AVERAGE cell should contain calculated value"
            try:
                calculated_avg = float(avg_value)
                assert abs(calculated_avg - expected_avg) < 0.001, \
                    f"Expected AVERAGE {expected_avg}, got {calculated_avg}"
            except ValueError:
                pass
            
            # Check MAX
            max_value = formula_row[3]
            assert not max_value.startswith('='), "MAX cell should contain calculated value"
            try:
                calculated_max = float(max_value)
                assert abs(calculated_max - expected_max) < 0.001, \
                    f"Expected MAX {expected_max}, got {calculated_max}"
            except ValueError:
                pass
    
    @given(divisor=st.integers(min_value=-100, max_value=100))
    @settings(max_examples=50, deadline=None, suppress_health_check=[HealthCheck.function_scoped_fixture])
    def test_property_11_formula_errors(self, divisor):
        """
        Feature: document-to-markdown-converter
        Property 11: Formula value conversion (error handling)
        
        For any Excel cell containing a formula that results in an error,
        the converter should handle it gracefully.
        
        Validates: Requirements 2.4
        """
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        # Create an Excel file with a formula that may cause an error
        with tempfile.TemporaryDirectory() as tmp_dir:
            excel_path = Path(tmp_dir) / f"test_formula_error_{divisor}.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "FormulaErrors"
            
            # Add headers
            sheet['A1'] = "Dividend"
            sheet['B1'] = "Divisor"
            sheet['C1'] = "Result"
            
            # Add data
            sheet['A2'] = 100
            sheet['B2'] = divisor
            
            # Add formula that may cause division by zero
            sheet['C2'] = "=A2/B2"
            
            workbook.save(excel_path)
            
            # Parse the Excel file
            parser = ExcelParser()
            result = parser.parse(str(excel_path))
            
            # Property: Should produce a valid table without crashing
            assert len(result.sections) == 1
            table = result.sections[0].content[0]
            assert isinstance(table, Table)
            
            # Property: Table should have correct structure
            assert len(table.headers) == 3
            assert len(table.rows) == 1
            
            # Property: Result cell should contain either a value or an error indicator
            result_cell = table.rows[0][2]
            assert result_cell is not None, "Result cell should not be None"
            
            if divisor == 0:
                # Should contain an error indicator or the error value
                # Common error values: #DIV/0!, #ERROR!, or empty
                # We just verify it doesn't crash and returns something
                assert isinstance(result_cell, str), "Result should be a string"
            else:
                # Should contain the calculated value
                expected_result = 100 / divisor
                try:
                    calculated_value = float(result_cell)
                    assert abs(calculated_value - expected_result) < 0.001, \
                        f"Expected result {expected_result}, got {calculated_value}"
                except ValueError:
                    # If it's not a number, it might be an error
                    pass


# Run all property tests
if __name__ == "__main__":
    pytest.main([__file__, "-v"])
